#!/usr/bin/env python3
"""
Pino Market Dashboard Updater (GitHub Actions friendly)

Updates two sheets in the Excel file:
- Dashboard (sector leaders + sector heat)
- Watchlist (any tickers you add)

Data sources:
- Stooq (free, no key): daily price history used to compute SMA/RSI/spreads
- Optional yfinance: dividend yield (best-effort)

Run locally:
  python pino_dashboard_updater.py --input Pino_Market_Dashboard_Auto.xlsx --output Pino_Market_Dashboard_Auto.xlsx
"""

from __future__ import annotations

import argparse
import io
from dataclasses import dataclass
from typing import Optional, Dict, List, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

LOOKBACK_DAYS = 320  # buffer for 200-day SMA

# -------------------------
# Heat + action rules
# -------------------------

def sector_heat(spread_pct: float) -> str:
    if spread_pct >= 7:
        return "🟢🟢🟢"
    if spread_pct >= 4:
        return "🟢🟢"
    if spread_pct >= 2:
        return "🟡🟡"
    if spread_pct >= 0:
        return "🟡"
    if spread_pct >= -2:
        return "🟠"
    return "🔴"


def action_label(
    score: Optional[float],
    stock_spread_50_200: Optional[float],
    price_vs_50: Optional[float],
    sector_spread_50_200: Optional[float],
) -> str:
    if stock_spread_50_200 is None or price_vs_50 is None:
        return "WAIT (missing data)"

    # Sector spread is optional on Watchlist if user leaves sector blank.
    # If missing, we just don't require sector filters.
    sector_ok = True if sector_spread_50_200 is None else (sector_spread_50_200 >= 4)

    score_ok = True if score is None else (score >= 80)

    # Ideal "safe add" setup
    if (
        score_ok
        and 3 <= stock_spread_50_200 <= 6
        and -5 <= price_vs_50 <= -2
        and sector_ok
    ):
        return "ADD (ideal dip)"

    # Softer add rule
    if score_ok and stock_spread_50_200 > 0 and -6 <= price_vs_50 <= 0:
        return "ADD (on dip)"

    if stock_spread_50_200 > 0:
        return "HOLD / WAIT FOR DIP"

    return "AVOID / WAIT (trend weak)"


# -------------------------
# Technical calculations
# -------------------------

def sma(series: pd.Series, window: int) -> pd.Series:
    return series.rolling(window=window, min_periods=window).mean()


def rsi(close: pd.Series, period: int = 14) -> pd.Series:
    delta = close.diff()
    gain = delta.where(delta > 0, 0.0)
    loss = -delta.where(delta < 0, 0.0)
    avg_gain = gain.ewm(alpha=1 / period, min_periods=period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1 / period, min_periods=period, adjust=False).mean()
    rs = avg_gain / avg_loss.replace(0, pd.NA)
    return 100 - (100 / (1 + rs))


# -------------------------
# Data fetching
# -------------------------

def fetch_stooq_daily(symbol: str) -> pd.DataFrame:
    sym = symbol.strip().lower()
    if "." not in sym:
        sym = f"{sym}.us"
    url = f"https://stooq.com/q/d/l/?s={sym}&i=d"
    r = requests.get(url, timeout=25)
    r.raise_for_status()
    df = pd.read_csv(io.StringIO(r.text))
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.sort_values("Date")
    return df


def try_dividend_yield_yfinance(symbol: str) -> Optional[float]:
    try:
        import yfinance as yf  # optional
        t = yf.Ticker(symbol)
        info = getattr(t, "info", {}) or {}
        dy = info.get("dividendYield", None)
        if dy is None:
            return None
        return float(dy) * 100.0
    except Exception:
        return None


@dataclass
class Metrics:
    sma50: Optional[float]
    sma200: Optional[float]
    spread_50_200: Optional[float]
    last_price: Optional[float]
    price_vs_50: Optional[float]
    rsi14: Optional[float]
    div_yield: Optional[float]


def compute_metrics(symbol: str) -> Metrics:
    div_yield = try_dividend_yield_yfinance(symbol)
    try:
        df = fetch_stooq_daily(symbol)
    except Exception:
        return Metrics(None, None, None, None, None, None, div_yield)

    close = df["Close"].astype(float)
    if len(close) < 210:
        last = float(close.iloc[-1]) if len(close) else None
        return Metrics(None, None, None, last, None, None, div_yield)

    df = df.tail(LOOKBACK_DAYS)
    close = df["Close"].astype(float)

    s50 = sma(close, 50)
    s200 = sma(close, 200)
    r = rsi(close, 14)

    last_price = float(close.iloc[-1])
    sma50_last = float(s50.iloc[-1]) if pd.notna(s50.iloc[-1]) else None
    sma200_last = float(s200.iloc[-1]) if pd.notna(s200.iloc[-1]) else None
    rsi_last = float(r.iloc[-1]) if pd.notna(r.iloc[-1]) else None

    spread = None
    if sma50_last is not None and sma200_last is not None and sma200_last != 0:
        spread = (sma50_last - sma200_last) / sma200_last * 100.0

    pvs50 = None
    if sma50_last is not None and sma50_last != 0:
        pvs50 = (last_price - sma50_last) / sma50_last * 100.0

    return Metrics(sma50_last, sma200_last, spread, last_price, pvs50, rsi_last, div_yield)


# -------------------------
# Excel styling
# -------------------------

FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def heat_fill(heat: str):
    if heat in ("🟢🟢🟢", "🟢🟢"):
        return FILL_GREEN
    if heat in ("🟡🟡", "🟡"):
        return FILL_YELLOW
    if heat == "🟠":
        return FILL_ORANGE
    return FILL_RED


def fmt(x: Optional[float], nd: int = 2) -> Optional[float]:
    if x is None:
        return None
    return round(float(x), nd)


def build_col_map(ws):
    headers = [c.value for c in ws[1]]
    return {h: i + 1 for i, h in enumerate(headers)}


def update_dashboard_sheet(wb):
    ws = wb["Dashboard"]
    col = build_col_map(ws)

    # Compute sector ETF spreads
    sector_spread_by_etf: Dict[str, Optional[float]] = {}
    for row in range(2, ws.max_row + 1):
        etf = ws.cell(row=row, column=col["Sector ETF (for heat)"]).value
        if not etf:
            continue
        etf = str(etf).strip().upper()
        if etf in sector_spread_by_etf:
            continue
        sector_spread_by_etf[etf] = compute_metrics(etf).spread_50_200

    # Update rows
    for row in range(2, ws.max_row + 1):
        etf = ws.cell(row=row, column=col["Sector ETF (for heat)"]).value
        ticker = ws.cell(row=row, column=col["Ticker"]).value
        score_cell = ws.cell(row=row, column=col["Score (0-100) [optional]"]).value

        etf = str(etf).strip().upper() if etf else None
        ticker = str(ticker).strip().upper() if ticker else None
        sector_spread = sector_spread_by_etf.get(etf, None) if etf else None

        # Sector
        if sector_spread is not None:
            ws.cell(row=row, column=col["Sector ETF 50/200 Spread %"]).value = fmt(sector_spread, 2)
            heat = sector_heat(sector_spread)
            hc = ws.cell(row=row, column=col["Sector Heat"])
            hc.value = heat
            hc.fill = heat_fill(heat)
        else:
            ws.cell(row=row, column=col["Sector ETF 50/200 Spread %"]).value = None
            hc = ws.cell(row=row, column=col["Sector Heat"])
            hc.value = "?"
            hc.fill = FILL_YELLOW

        if not ticker:
            continue

        m = compute_metrics(ticker)
        ws.cell(row=row, column=col["50d SMA"]).value = fmt(m.sma50, 2)
        ws.cell(row=row, column=col["200d SMA"]).value = fmt(m.sma200, 2)
        ws.cell(row=row, column=col["50/200 Spread %"]).value = fmt(m.spread_50_200, 2)
        ws.cell(row=row, column=col["Last Price"]).value = fmt(m.last_price, 2)
        ws.cell(row=row, column=col["Price vs 50d %"]).value = fmt(m.price_vs_50, 2)
        ws.cell(row=row, column=col["RSI (14)"]).value = fmt(m.rsi14, 1)
        ws.cell(row=row, column=col["Dividend Yield [optional]"]).value = fmt(m.div_yield, 2)

        # Score
        score = None
        try:
            score = float(score_cell) if score_cell not in (None, "") else None
        except Exception:
            score = None

        act = action_label(score, m.spread_50_200, m.price_vs_50, sector_spread)
        ws.cell(row=row, column=col["Action"]).value = act


def update_watchlist_sheet(wb):
    if "Watchlist" not in wb.sheetnames:
        return
    ws = wb["Watchlist"]
    col = build_col_map(ws)

    # Cache sector ETF spreads used in watchlist
    sector_spread_by_etf: Dict[str, Optional[float]] = {}

    for row in range(2, ws.max_row + 1):
        ticker = ws.cell(row=row, column=col["Ticker"]).value
        if not ticker or str(ticker).strip() == "":
            continue
        ticker = str(ticker).strip().upper()

        sector_etf = ws.cell(row=row, column=col["Sector ETF [optional]"]).value
        sector_etf = str(sector_etf).strip().upper() if sector_etf else None

        # Sector ETF spread (optional)
        sector_spread = None
        if sector_etf:
            if sector_etf not in sector_spread_by_etf:
                sector_spread_by_etf[sector_etf] = compute_metrics(sector_etf).spread_50_200
            sector_spread = sector_spread_by_etf.get(sector_etf)

        # Stock metrics
        m = compute_metrics(ticker)
        ws.cell(row=row, column=col["50d SMA"]).value = fmt(m.sma50, 2)
        ws.cell(row=row, column=col["200d SMA"]).value = fmt(m.sma200, 2)
        ws.cell(row=row, column=col["50/200 Spread %"]).value = fmt(m.spread_50_200, 2)
        ws.cell(row=row, column=col["Last Price"]).value = fmt(m.last_price, 2)
        ws.cell(row=row, column=col["Price vs 50d %"]).value = fmt(m.price_vs_50, 2)
        ws.cell(row=row, column=col["RSI (14)"]).value = fmt(m.rsi14, 1)
        ws.cell(row=row, column=col["Dividend Yield [optional]"]).value = fmt(m.div_yield, 2)

        if sector_spread is not None:
            ws.cell(row=row, column=col["Sector ETF 50/200 Spread %"]).value = fmt(sector_spread, 2)
            heat = sector_heat(sector_spread)
            hc = ws.cell(row=row, column=col["Sector Heat"])
            hc.value = heat
            hc.fill = heat_fill(heat)
        else:
            ws.cell(row=row, column=col["Sector ETF 50/200 Spread %"]).value = None
            hc = ws.cell(row=row, column=col["Sector Heat"])
            hc.value = ""
            hc.fill = PatternFill()  # reset

        # Score
        score_cell = ws.cell(row=row, column=col["Score (0-100) [optional]"]).value
        score = None
        try:
            score = float(score_cell) if score_cell not in (None, "") else None
        except Exception:
            score = None

        ws.cell(row=row, column=col["Action"]).value = action_label(score, m.spread_50_200, m.price_vs_50, sector_spread)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    wb = load_workbook(args.input)
    if "Dashboard" not in wb.sheetnames:
        raise SystemExit("Dashboard sheet not found.")

    update_dashboard_sheet(wb)
    update_watchlist_sheet(wb)

    wb.save(args.output)
    print(f"Updated dashboard saved to: {args.output}")


if __name__ == "__main__":
    main()
